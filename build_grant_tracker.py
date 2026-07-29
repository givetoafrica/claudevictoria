#!/usr/bin/env python3
"""Render grant-data/grants.json into the Give to Africa grant pipeline workbook.

The JSON file is the source of truth. This script only formats it, so the daily
monitor edits JSON and re-runs this rather than mutating the .xlsx in place.

    python3 build_grant_tracker.py
"""

import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "grant-data", "grants.json")
OUT = os.path.join(HERE, "Give-to-Africa-Grant-Pipeline-2026.xlsx")

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
BODY_TOP = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BANDS = {
    1: ("1 - Time-sensitive", "FCE4D6"),
    2: ("2 - Open / rolling", "E2EFDA"),
    3: ("3 - Calendar for next cycle", "FFF2CC"),
    4: ("4 - Low priority", "F2F2F2"),
}

COLUMNS = [
    ("Priority", "priority", 17),
    ("Grant / Funder Name", "name", 34),
    ("Funder Type", "funder_type", 20),
    ("Status", "status", 20),
    ("Amount or Ceiling", "amount", 26),
    ("Deadline", "deadline", 20),
    ("Deadline (sortable)", "deadline_iso", 15),
    ("Eligibility Notes", "eligibility", 40),
    ("Program Fit (Give to Africa)", "fit", 44),
    ("Website", "website", 26),
    ("Program Contact Email", "email", 26),
    ("Phone", "phone", 16),
    ("HOW TO APPLY", "how_to_apply", 52),
    ("Link (source checked)", "link", 40),
    ("Recommended Next Step", "next_step", 40),
    ("Verification", "verification", 30),
    ("First Seen", "first_seen", 13),
]


def load():
    with open(DATA, encoding="utf-8") as fh:
        return json.load(fh)


def sort_key(g):
    """Within a band, soonest real deadline first; undated entries last."""
    iso = g.get("deadline_iso", "unknown")
    dated = len(iso) == 10 and iso[4] == "-"
    return (g["priority"], 0 if dated else 1, iso if dated else "", g["name"])


def build():
    grants = sorted(load(), key=sort_key)
    wb = Workbook()

    ws = wb.active
    ws.title = "Grant Pipeline"

    title = ws.cell(row=1, column=1)
    title.value = (
        "Give to Africa - Grant Pipeline (Government, Multilateral & Foundation)  |  "
        f"Last updated {date.today():%B %d, %Y}  |  {len(grants)} opportunities tracked"
    )
    title.font = Font(name=FONT, size=12, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 22

    for col, (header, _, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col, value=header)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[2].height = 30

    for i, g in enumerate(grants, start=3):
        label, color = BANDS[g["priority"]]
        band = PatternFill("solid", fgColor=color)
        for col, (_, key, _) in enumerate(COLUMNS, start=1):
            val = label if key == "priority" else g.get(key, "")
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = BODY_TOP
            c.border = BORDER
            if key == "priority":
                c.fill = band
                c.font = Font(name=FONT, size=10, bold=True)
            elif key == "status":
                c.fill = band
            elif key == "verification" and str(val).startswith(("NEEDS VERIFICATION", "CORRECTED")):
                c.font = Font(name=FONT, size=10, bold=True, color="C00000")
            elif key == "how_to_apply":
                text = str(val)
                if text.startswith(("NO ", "CLOSED", "NOT A GRANT")):
                    c.font = Font(name=FONT, size=10, color="C00000")
                elif text.startswith("APPLY"):
                    c.font = Font(name=FONT, size=10, bold=True, color="1F6F3D")
        ws.row_dimensions[i].height = 96

    for col, (_, _, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(grants)}"

    build_notes(wb, grants)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(grants)} opportunities)")


def build_notes(wb, grants):
    ws = wb.create_sheet("How to Use + Method")

    def put(r, c, val, bold=False, size=10, color="000000", height=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT, size=size, bold=bold, color=color)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if height:
            ws.row_dimensions[r].height = height
        return cell

    put(1, 1, "How to use this tracker", bold=True, size=12, color="1F3864")

    r = 3
    put(r, 1, "Priority bands", bold=True)
    r += 1
    for pri, meaning in [
        (1, "Deadline inside the next ~75 days. Act on these first."),
        (2, "Open, rolling, or partnership-based - you can move whenever you are ready."),
        (3, "Closed now but recurring. Several of the best strategic fits live here - they need calendar reminders, not proposals."),
        (4, "Closed, ineligible, or poor fit. Kept so the same ground is not re-searched next month."),
    ]:
        label, color = BANDS[pri]
        put(r, 1, label, bold=True).fill = PatternFill("solid", fgColor=color)
        put(r, 2, meaning)
        r += 1

    r += 1
    put(r, 1, "Automated daily monitoring", bold=True)
    r += 1
    put(r, 2,
        "This workbook is regenerated by the 'grant-monitor' skill in .claude/skills/grant-monitor/. "
        "The real source of truth is grant-data/grants.json - edit that, then run build_grant_tracker.py. "
        "The monitor runs a rotating set of searches each weekday, so it surfaces genuinely new calls "
        "instead of re-reporting the same results daily. It also re-bands opportunities automatically as "
        "deadlines approach and pass.",
        height=74)
    r += 2

    put(r, 1, "Contact information policy", bold=True)
    r += 1
    put(r, 1, '"not publicly listed"')
    put(r, 2,
        "The email or phone was looked for and is genuinely not published on the funder's site or the "
        "official listing. No contact detail in this file was guessed or constructed - every address shown "
        "appears in a published source.",
        height=46)
    r += 2

    put(r, 1, "Verification column", bold=True)
    r += 1
    put(r, 1, '"confirmed"')
    put(r, 2, "Amount, deadline and eligibility were consistent across the sources checked.")
    r += 1
    put(r, 1, '"NEEDS VERIFICATION"', bold=True)
    put(r, 2,
        "Sources conflicted or the funder's own page could not be reached. Confirm directly with the funder "
        "before investing proposal time. Read the IFC Agritech row carefully - sites impersonating IFC grant "
        "programs are documented, so verify on ifc.org itself and never pay a fee to apply.",
        height=60)
    r += 2

    put(r, 1, "Counts", bold=True)
    r += 1
    put(r, 1, "Total opportunities tracked")
    ws.cell(row=r, column=2, value=len(grants)).font = Font(name=FONT, size=10, bold=True)
    r += 1
    for pri in (1, 2, 3, 4):
        label, _ = BANDS[pri]
        put(r, 1, label)
        ws.cell(row=r, column=2, value=sum(1 for g in grants if g["priority"] == pri)).font = BODY_FONT
        r += 1

    r += 1
    put(r, 1, "What the research has found so far", bold=True)
    r += 1
    for note in [
        "The U.S. federal pipeline for Africa development work is thin. Open federal calls are either geographically mismatched (DRL Track 2.0 targets Armenia/Azerbaijan, Cambodia/Thailand, DRC/Rwanda) or far above current scale (USDA Food for Progress at $28-35M per award, with Ghana off the FY2026 priority list).",
        "The realistic government money is embassy-level and partner-fronted: U.S. Embassy Accra Self-Help and Public Diplomacy, USADF, and Canada's CFLI. Three of those four need a Ghana-registered applicant, making MDF Ghana the applicant of record and Give to Africa the technical and compliance partner. That structure is worth formalizing before the next cycle.",
        "Canada's CFLI Ghana is the eligibility exception - it explicitly accepts international NGOs working on local development, so Give to Africa can apply as itself.",
        "The two highest-value finds are not U.S. sources. UNDP/GCF's Ghana Shea Landscape project (GSLERP) matches on geography, commodity, beneficiaries and climate framing at once, and actively seeks NGO partners who bring women's cooperatives. KfW's Investing for Employment covers up to 90% of costs for non-revenue non-profit projects at EUR 800K-10M scale.",
        "Non-government standouts: the Global Fund for Community Foundations takes rolling concept notes and funds exactly the community-philanthropy infrastructure YENDAA represents. AAK/Beiersdorf is a procurement relationship rather than a grant - a larger budget line than CSR if the cooperatives can supply traceable shea.",
    ]:
        put(r, 1, "-")
        put(r, 2, note, height=62)
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 108


if __name__ == "__main__":
    build()
